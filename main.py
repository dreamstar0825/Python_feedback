import pandas as pd
from openpyxl import load_workbook
from transformers import pipeline

# Load sentiment analysis model
model_path = 'distilbert-base-uncased-finetuned-sst-2-english'

classifier = pipeline('sentiment-analysis', model=model_path)

def read_and_classify(file_path):
    # Load the workbook and active sheet
    wb = load_workbook(filename=file_path, data_only=True)
    ws = wb.active

    # Dictionary to store results
    # Iterate through cells with yellow fill
    for row in ws.iter_rows(min_row=2):
        for idx, cell in enumerate(row):

            if idx % 2 !=0:
                feedback = str(cell.value)
                truncated_feedback = feedback[:512]
                result = classifier(truncated_feedback)
                ws.cell(row = cell.row, column = cell.column+1).value = result[0]['label']
    wb.save(file_path)
    return 'success'

feedback_results = read_and_classify('./LearnerExperience.xlsx')

print(feedback_results)